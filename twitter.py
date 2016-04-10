import requests
from bs4 import BeautifulSoup
import time
from openpyxl import load_workbook
from openpyxl import Workbook

country_name=[]
ind_url=[]
h_num=2
c_num=2
ind_num=2
nam_num = 2

def getting_country_name():   # this function will get country name from the locally saved file mainpage.txt and save it it required format in a list
    x=open('mainpage.txt','r')
    tree=x.read()
    x.close()
    y=[]

    soup=BeautifulSoup(tree,'lxml')
    for a in soup.findAll('article',{'class':'country'}):
        for b in a.findAll('form'):
            for c in b.findAll('select'):
                for d in c.findAll('option'):
                    y.append(d.string)
    for x in y:
        x=x.replace(" ",'-')
        x=x.replace(',','-')
        x=x.replace('.','-')
        x=x.replace('--','-')
        y=list(x)
        if (y[-1]=='-'):   # checking if the last element is "-" or not and if it is then removing it
           y[-1]="/"
        else :
            y.append("/")
        x="".join(y)
        country_name.append(x)
def page_empty_check(soup):  # this function will check if the page is empty or not
    count=0
    for a in soup.findAll('div',{'class':'brand-table-placeholder'}):  # these loops check if the page is empty or not
        for b in a.findAll('table',{'class':'brand-table-list'}):
            for c in b.findAll('tr'):
                count+=1
                if (count == 2):
                   return count
    return count
def wrong_page_check (soup): # this function checks for landing in the wrong page i.e. Whopsi error 404 page not found one
    for a in soup.findAll('title'):
        if "404" in a.string :
            return 1
        else :
            return 0
def check_show_more_button(soup): # this check if show more button is present in the page or not
    for a in soup.findAll('div',{'class':'more-center-link'}):
         for b in a.findAll('a'):
             if "Show More" in b.string :
                return 1
    return 0


def main(temp_url,country,ind):
    page_num_start=1
    page_num_end = 5
    while True :
        url = temp_url+"page-"+str(page_num_start)+"-"+str(page_num_end)+"/"
        print (url)
        while True :
            try :
                #time.sleep(10)
                source_code=requests.get(url)  #getting page source
                break
            except requests.exceptions.ConnectionError:  # to handle error if website refuses connection due to multiple retries
                print("Wait !! Website Refusing Connection.... ")
                time.sleep(30)
        if country.lower() not in source_code.url :   # this checks if the url has been redirected to another webpage without country or not and if yes then it stops the loop
            break

        text_source=source_code.text #converting it to text file
        soup=BeautifulSoup(text_source,'html.parser')  # making soup object
        flag = wrong_page_check(soup)
        if flag == 1 :
            break
        count = page_empty_check(soup)
        if (count != 1) :   # if this condition is true this implies everything is fine and now getting the handles
            for a in soup.findAll('div',{'class','item'}):
                for b in a.findAll('h2'):
                    for c in b.findAll('span'):
                        temp_var=c.string
                        twitter_handle = (temp_var[temp_var.find("(")+1:temp_var.find(")")])
                        name = (temp_var[:temp_var.find("(")])
                       # print(name+" - "+ twitter_handle + " - " + country[:-1] + " - " + ind[29:-1])
                        writing_to_file(name , twitter_handle,country[:-1],ind[29:-1])
        flag2=check_show_more_button(soup)
        if flag2 == 1 :
            pass
        else :
            break
        page_num_start+=5
        page_num_end += 5
def getting_industry_link():
    from bs4 import BeautifulSoup
    file_obj = open('mainpage.txt','r')
    text = file_obj.read()
    file_obj.close()
    y=[]  # to temporary store list of indusrty urls
    soup = BeautifulSoup(text,'html.parser')
    for a in soup.findAll('div'):
        for b in a.findAll('ul',{'class':'multi-dropdown-list'}):
            for c in b.findAll('li'):
                for d in c.findAll ('a'):
                    y.append(d.get('href'))
    z=set(y)
    for x in z:
        ind_url.append(x)
def changing_url(country,industry): # this function will convert the urls in a crawlerable format
    temp_list= list(industry)
    temp_country_list=list(country)
    temp_list[29:29]=temp_country_list
    temp_list[0:0]="http://www.socialbakers.com"
    x="".join(temp_list)
    return (x.lower())
def writing_to_file(name , handle,country,industry):
    global h_num , c_num ,ind_num ,nam_num
    wb = load_workbook('twitter.xlsx')
    ws = wb.active
    ws.cell(row = h_num, column = 2).value=handle
    ws.cell(row = c_num, column = 3).value=country
    ws.cell(row = ind_num, column = 4).value=industry
    ws.cell(row = nam_num, column = 1).value=name
    h_num+=1
    c_num+=1
    ind_num+=1
    nam_num+=1
    wb.save('twitter.xlsx')

getting_country_name()
getting_industry_link()
wb = Workbook() # creating excel file
ws = wb.active
ws.cell(row =1, column = 2).value="Twitter Handle"
ws.cell(row =1, column = 3).value="Country"
ws.cell(row =1, column = 4).value="Industry"
ws.cell(row =1, column = 1).value="Name"
wb.save('twitter.xlsx') # saving the file
del country_name[0]  # because first one wil be that Select Country
for temp_country in country_name :
    for temp_industry_url in ind_url :
        url = changing_url(temp_country,temp_industry_url)
        main(url,temp_country,temp_industry_url)





